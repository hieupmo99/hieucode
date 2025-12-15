import json
import time
import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill, Alignment, Font

class TikTokAPIDocScraperSelenium:
    def __init__(self, headless=True):
        """
        Initialize Selenium-based scraper for JavaScript-rendered content
        """
        self.base_url = "https://business-api.tiktok.com"
        self.driver = None
        self.headless = headless
        
    def setup_driver(self):
        """
        Setup Chrome driver with appropriate options
        """
        chrome_options = Options()
        if self.headless:
            chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        # Try common ChromeDriver paths
        chromedriver_paths = [
            "/opt/homebrew/bin/chromedriver",  # Homebrew on Apple Silicon
            "/usr/local/bin/chromedriver",      # Homebrew on Intel Mac
            "chromedriver"                      # System PATH
        ]
        
        driver_service = None
        for path in chromedriver_paths:
            try:
                driver_service = Service(path)
                self.driver = webdriver.Chrome(
                    service=driver_service, options=chrome_options
                )
                print(f"✓ Using ChromeDriver at: {path}\n")
                break
            except Exception:
                continue
        
        # Fallback to default if none of the paths work
        if not self.driver:
            try:
                self.driver = webdriver.Chrome(options=chrome_options)
                print("✓ Using ChromeDriver from system PATH\n")
            except Exception as e:
                print(f"✗ Error: Could not initialize ChromeDriver")
                print(f"  {e}")
                print("\nPlease ensure ChromeDriver is installed:")
                print("  brew install chromedriver")
                raise
        
        self.driver.set_page_load_timeout(30)
        
    def get_iframe_content(self, url, wait_time=10, max_retries=3):
        """
        Load page and extract iframe content using Selenium with retry logic
        """
        for attempt in range(max_retries):
            try:
                print(f"Loading page: {url}")
                
                # Check if driver is still alive
                try:
                    _ = self.driver.current_url
                except Exception:
                    print("  ⚠ Driver session lost, recreating...")
                    self.driver.quit()
                    self.setup_driver()
                
                self.driver.get(url)
                
                # Wait for iframe to load
                print("Waiting for iframe to load...")
                iframe = WebDriverWait(self.driver, wait_time).until(
                    EC.presence_of_element_located((By.TAG_NAME, "iframe"))
                )
                
                print("✓ Iframe found!")
                
                # Get iframe src
                iframe_src = iframe.get_attribute('src')
                print(f"Iframe URL: {iframe_src[:100]}...")
                
                # Switch to iframe
                self.driver.switch_to.frame(iframe)
                
                # Wait for content to load
                time.sleep(2)
                
                # Get page source from iframe
                iframe_html = self.driver.page_source
                
                # Switch back to main content
                self.driver.switch_to.default_content()
                
                return iframe_html, iframe_src
                
            except Exception as e:
                print(f"  ⚠ Attempt {attempt + 1}/{max_retries} failed: {str(e)[:100]}")
                
                if attempt < max_retries - 1:
                    print(f"  → Retrying in 3 seconds...")
                    time.sleep(3)
                    
                    # Try to recover the driver
                    try:
                        self.driver.switch_to.default_content()
                    except:
                        pass
                    
                    try:
                        self.driver.quit()
                    except:
                        pass
                    
                    self.setup_driver()
                else:
                    print(f"  ✗ All retry attempts failed")
                    return None, None
        
        return None, None
    
    def get_endpoint_links_from_page(self, page_url):
        """
        Extract all endpoint links from the documentation page
        """
        print("\n" + "="*80)
        print("STEP 1: Getting list of endpoint documentation pages")
        print("="*80)
        print(f"\nFetching endpoint links from: {page_url}")
        
        iframe_html, iframe_src = self.get_iframe_content(page_url)
        
        if not iframe_html:
            print("✗ Could not load iframe content")
            return []
        
        soup = BeautifulSoup(iframe_html, 'html.parser')
        
        endpoint_links = []
        all_links = soup.find_all('a', href=True)
        
        print(f"Found {len(all_links)} total links in iframe")
        
        for link in all_links:
            href = link.get('href')
            text = link.get_text(strip=True)
            
            # Look for documentation links with id parameter
            if href and 'id=' in href and text and len(text) > 3:
                # Convert to full URL
                if href.startswith('/portal/docs'):
                    full_url = self.base_url + href
                elif not href.startswith('http'):
                    full_url = self.base_url + '/portal/docs' + href
                else:
                    full_url = href
                
                endpoint_links.append({
                    'text': text,
                    'url': full_url
                })
        
        # Remove duplicates
        seen = set()
        unique_links = []
        for link in endpoint_links:
            if link['url'] not in seen:
                seen.add(link['url'])
                unique_links.append(link)
        
        print(f"✓ Found {len(unique_links)} unique endpoint links\n")
        return unique_links
    
    def extract_supported_fields(self, description_text):
        """
        Extract supported field names from description text
        Looks for "Supported fields:" section and extracts field names (before colon or first word)
        Example: "is_business_account：Whether..." -> "is_business_account"
        """
        if not description_text:
            return []
        
        supported_fields = []
        
        # Look for "Supported fields:" or "Supported fields" pattern
        import re
        # Pattern to find "Supported fields:" section
        pattern = r'(?:Supported\s+fields|supported\s+fields)[:：]?\s*(.*?)(?:\n\n|\Z)'
        match = re.search(pattern, description_text, re.IGNORECASE | re.DOTALL)
        
        if match:
            fields_section = match.group(1).strip()
            
            # Split by newlines or bullet points
            # Handle different formats:
            # 1. Bullet points: "• field_name: description"
            # 2. Numbered: "1. field_name: description"
            # 3. Plain list: "field_name: description\nfield_name2: description"
            # 4. Chinese colon: "field_name：description"
            
            # Split by newlines first
            lines = re.split(r'\n+', fields_section)
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Remove bullet points (•, -, *, etc.)
                line = re.sub(r'^[•\-\*\d+\.\)]\s*', '', line)
                
                # Extract field name (everything before colon or first word)
                # Handle both English colon (:) and Chinese colon (：)
                colon_match = re.match(r'^([a-zA-Z_][a-zA-Z0-9_]*)[：:]', line)
                if colon_match:
                    field_name = colon_match.group(1)
                    supported_fields.append(field_name)
                else:
                    # If no colon, try to extract first word (field name)
                    word_match = re.match(r'^([a-zA-Z_][a-zA-Z0-9_]*)', line)
                    if word_match:
                        field_name = word_match.group(1)
                        supported_fields.append(field_name)
        
        return supported_fields
    
    def extract_table_data(self, table):
        """
        Extract data from a table with improved parsing
        Also extracts supported fields from parameter descriptions
        """
        if not table:
            return []
        
        rows = table.find_all('tr')
        data = []
        
        # Check if first row is header row
        first_row_cells = rows[0].find_all(['td', 'th']) if rows else []
        is_header_row = any('th' in str(cell) for cell in first_row_cells) or len(first_row_cells) == 0
        
        start_idx = 1 if is_header_row else 0
        
        for row in rows[start_idx:]:
            cells = row.find_all(['td', 'th'])
            
            # Skip empty rows
            if not cells:
                continue
            
            # Clean cell text
            cell_texts = [cell.get_text(strip=True) for cell in cells]
            
            # Handle different table structures
            if len(cells) >= 3:
                # Standard 3-column: Field | Data Type | Description
                field = cell_texts[0]
                data_type = cell_texts[1] if len(cell_texts) > 1 else ''
                description = cell_texts[2] if len(cell_texts) > 2 else ''
                
                # Clean field name - remove "Required" suffix and extra whitespace
                if field:
                    field = field.replace('Required', '').replace('required', '').strip()
                    # Remove trailing punctuation/spaces
                    field = field.rstrip('.,;: ')
                
                # Extract supported fields from description
                supported_fields = self.extract_supported_fields(description)
                
                # Skip if field is empty or looks like a header row
                if field and field.lower() not in ['field', 'name', 'parameter', 'header', 'type', 'description']:
                    param_data = {
                        'field': field,
                        'data_type': data_type,
                        'description': description
                    }
                    if supported_fields:
                        param_data['supported_fields'] = supported_fields
                    data.append(param_data)
            elif len(cells) == 2:
                # 2-column: Field | Description or Name | Type
                field = cell_texts[0]
                second_col = cell_texts[1] if len(cell_texts) > 1 else ''
                
                # Clean field name
                if field:
                    field = field.replace('Required', '').replace('required', '').strip()
                    field = field.rstrip('.,;: ')
                
                # Try to detect if second column is type or description
                if field and field.lower() not in ['field', 'name', 'parameter', 'header', 'type', 'description']:
                    # Check if second column looks like a type (short, common types)
                    if second_col.lower() in ['string', 'integer', 'int', 'number', 'boolean', 'bool', 'array', 'object']:
                        param_data = {
                            'field': field,
                            'data_type': second_col,
                            'description': ''
                        }
                        data.append(param_data)
                    else:
                        # Second column is description - extract supported fields
                        supported_fields = self.extract_supported_fields(second_col)
                        param_data = {
                            'field': field,
                            'data_type': '',
                            'description': second_col
                        }
                        if supported_fields:
                            param_data['supported_fields'] = supported_fields
                        data.append(param_data)
            elif len(cells) == 1:
                # Single column - might be a list
                field = cell_texts[0]
                if field and field.lower() not in ['field', 'name', 'parameter', 'header']:
                    data.append({
                        'field': field,
                        'data_type': '',
                        'description': ''
                    })
        
        return data
    
    def extract_endpoint_details(self, url, endpoint_text):
        """
        Extract endpoint details from a documentation page
        """
        print(f"  Scraping details...")
        
        endpoint_data = {
            'endpoint_name': endpoint_text,
            'doc_url': url,
            'page_title': None,  # Will extract from page
            'endpoint': None,
            'method': None,
            'headers': [],
            'parameters': [],
            'response': [],
            'scrape_status': 'failed',  # Will be updated to 'success' if scraping works
            'error_message': None
        }
        
        try:
            iframe_html, iframe_src = self.get_iframe_content(url)
            
            if not iframe_html:
                print(f"  ✗ Could not load iframe")
                endpoint_data['error_message'] = 'Failed to load iframe after retries'
                return endpoint_data
            
            soup = BeautifulSoup(iframe_html, 'html.parser')
            
            # Extract page title - prioritize DefaultView_title__3-BU8 div
            title_div = soup.find('div', class_='DefaultView_title__3-BU8')
            if title_div:
                endpoint_data['page_title'] = title_div.get_text(strip=True)
                print(f"    ✓ Found title from DefaultView_title div")
            else:
                # Fallback to title tag
                title_tag = soup.find('title')
                if title_tag:
                    endpoint_data['page_title'] = title_tag.get_text(strip=True)
                else:
                    # Try to find h1 tag as title
                    h1_tag = soup.find('h1')
                    if h1_tag:
                        endpoint_data['page_title'] = h1_tag.get_text(strip=True)
                    else:
                        endpoint_data['page_title'] = endpoint_text  # Fallback to endpoint name
            
            # Extract endpoint URL
            text_content = soup.get_text()
            
            # Look for API endpoint URL
            api_url_pattern = r'https://business-api\.tiktok\.com/open_api/[v\d\./a-zA-Z_]+'
            matches = re.findall(api_url_pattern, text_content)
            if matches:
                endpoint_data['endpoint'] = matches[0]
            
            if not endpoint_data['endpoint']:
                rel_url_pattern = r'/open_api/[v\d\./a-zA-Z_]+'
                matches = re.findall(rel_url_pattern, text_content)
                if matches:
                    endpoint_data['endpoint'] = 'https://business-api.tiktok.com' + matches[0]
            
            # Extract HTTP method
            method_keywords = ['POST', 'GET', 'PUT', 'DELETE', 'PATCH']
            for method in method_keywords:
                if re.search(rf'\b{method}\b', text_content, re.IGNORECASE):
                    endpoint_data['method'] = method
                    break
            
            # Extract tables - improved detection for headers/parameters
            # Method 1: Look for headings followed by tables (more precise matching)
            all_headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'strong', 'b', 'div'])
            
            for heading in all_headings:
                heading_text = heading.get_text(strip=True).lower()
                
                # Skip if heading is too generic or empty
                if len(heading_text) < 3:
                    continue
                
                # Find next table (could be sibling or nested)
                next_table = heading.find_next('table')
                
                if not next_table:
                    continue
                
                # More precise header detection - must contain "header" keyword AND header-like fields
                header_keywords = ['header', 'request header', 'http header', 'request headers']
                if any(keyword in heading_text for keyword in header_keywords) and not endpoint_data['headers']:
                    table_data = self.extract_table_data(next_table)
                    if table_data:
                        # Check if table contains typical header fields (strict check)
                        header_field_patterns = ['access-token', 'content-type', 'authorization', 'x-', 'access_token', 'content_type']
                        has_header_fields = any(
                            any(hf in field.get('field', '').lower() for hf in header_field_patterns)
                            for field in table_data
                        )
                        # Also check field names don't look like parameters (video_id, business_id, etc.)
                        param_like_patterns = ['_id', 'video', 'business', 'text', 'page', 'limit']
                        has_param_like_fields = any(
                            any(pl in field.get('field', '').lower() for pl in param_like_patterns)
                            for field in table_data
                        )
                        
                        if has_header_fields and not has_param_like_fields:
                            endpoint_data['headers'] = table_data
                            print(f"    ✓ Extracted {len(endpoint_data['headers'])} headers")
                        elif has_header_fields:
                            # Has header fields but also param-like - might be mixed, filter it
                            filtered_headers = [
                                field for field in table_data
                                if any(hf in field.get('field', '').lower() for hf in header_field_patterns)
                            ]
                            if filtered_headers:
                                endpoint_data['headers'] = filtered_headers
                                print(f"    ✓ Extracted {len(endpoint_data['headers'])} headers (filtered)")
                
                # More precise parameter detection - exclude headers and responses
                param_keywords = ['parameter', 'request parameter', 'request parameters', 'query parameter', 
                                 'body parameter', 'path parameter', 'parameters', 'request body']
                response_keywords = ['response', 'response field', 'response body', 'response parameter']
                
                # Check if this is a parameter section (not header, not response)
                is_param_section = (any(keyword in heading_text for keyword in param_keywords) and 
                                   not any(keyword in heading_text for keyword in response_keywords) and
                                   'header' not in heading_text.lower())
                
                if is_param_section and not endpoint_data['parameters']:
                    table_data = self.extract_table_data(next_table)
                    if table_data:
                        # Filter out header fields and non-parameter fields from parameters
                        header_fields = ['access-token', 'content-type', 'authorization']
                        # Also filter out fields that look like metadata (endpoint path, etc.)
                        metadata_patterns = ['endpoint', 'path', 'url', 'method', 'response', 'status']
                        
                        filtered_data = []
                        for field in table_data:
                            field_name_lower = field.get('field', '').lower()
                            # Skip header fields
                            if any(hf in field_name_lower for hf in header_fields):
                                continue
                            # Skip metadata fields
                            if any(mp in field_name_lower for mp in metadata_patterns):
                                continue
                            # Skip if field name looks like a heading/description rather than a parameter
                            if field_name_lower in ['new', 'parameter', 'field', 'name', 'type', 'description']:
                                continue
                            filtered_data.append(field)
                        
                        if filtered_data:
                            endpoint_data['parameters'] = filtered_data
                            print(f"    ✓ Extracted {len(endpoint_data['parameters'])} parameters")
                        elif table_data:
                            # If all were filtered, might be wrong table - but keep it
                            endpoint_data['parameters'] = table_data
                            print(f"    ✓ Extracted {len(endpoint_data['parameters'])} parameters (note: may include headers)")
                
                # Response detection
                if any(keyword in heading_text for keyword in response_keywords) and not endpoint_data['response']:
                    endpoint_data['response'] = self.extract_table_data(next_table)
                    if endpoint_data['response']:
                        print(f"    ✓ Extracted {len(endpoint_data['response'])} response fields")
            
            # Method 2: If still not found, look for all tables and check their context
            if not endpoint_data['headers'] or not endpoint_data['parameters']:
                all_tables = soup.find_all('table')
                for table in all_tables:
                    # Check preceding text/headings for context
                    prev_text = ''
                    prev_elem = table.find_previous(['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'strong', 'b', 'p'])
                    if prev_elem:
                        prev_text = prev_elem.get_text(strip=True).lower()
                    
                    # Check table headers for clues
                    table_headers = table.find_all(['th'])
                    header_text = ' '.join([th.get_text(strip=True).lower() for th in table_headers[:3]])
                    
                    combined_text = (prev_text + ' ' + header_text).lower()
                    
                    # Try to identify header table
                    if not endpoint_data['headers'] and ('header' in combined_text or 
                                                         any(h in combined_text for h in ['access-token', 'content-type', 'authorization'])):
                        endpoint_data['headers'] = self.extract_table_data(table)
                        if endpoint_data['headers']:
                            print(f"    ✓ Extracted {len(endpoint_data['headers'])} headers (from context)")
                    
                    # Try to identify parameter table
                    elif not endpoint_data['parameters'] and ('parameter' in combined_text or 
                                                             'field' in header_text or 'name' in header_text):
                        # Make sure it's not a response table
                        if 'response' not in combined_text:
                            endpoint_data['parameters'] = self.extract_table_data(table)
                            if endpoint_data['parameters']:
                                print(f"    ✓ Extracted {len(endpoint_data['parameters'])} parameters (from context)")
                    
                    # Try to identify response table
                    elif not endpoint_data['response'] and 'response' in combined_text:
                        endpoint_data['response'] = self.extract_table_data(table)
                        if endpoint_data['response']:
                            print(f"    ✓ Extracted {len(endpoint_data['response'])} response fields (from context)")
            
            # Extract example code blocks (curl, python, etc.)
            example_code = []
            
            # Method 1: Look for <pre><code> blocks
            pre_blocks = soup.find_all('pre')
            for pre_block in pre_blocks:
                code_elem = pre_block.find('code')
                if code_elem:
                    code_text = code_elem.get_text()
                else:
                    code_text = pre_block.get_text()
                
                code_text_clean = code_text.strip()
                
                # Look for meaningful code examples (lower threshold for detection)
                if len(code_text_clean) > 20:
                    # Detect language from class or content
                    language = 'unknown'
                    
                    # Check pre block classes
                    pre_classes = pre_block.get('class', [])
                    for cls in pre_classes:
                        cls_str = str(cls).lower()
                        if 'language' in cls_str or 'lang' in cls_str:
                            language = cls_str.split('-')[-1] if '-' in cls_str else cls_str
                            break
                    
                    # Check code element classes
                    if code_elem:
                        code_classes = code_elem.get('class', [])
                        for cls in code_classes:
                            cls_str = str(cls).lower()
                            if 'language' in cls_str or 'lang' in cls_str:
                                language = cls_str.split('-')[-1] if '-' in cls_str else cls_str
                                break
                    
                    # Detect from content
                    code_lower = code_text_clean.lower()
                    if 'curl' in code_lower or code_lower.startswith('curl'):
                        language = 'curl' if language == 'unknown' else language
                    elif 'import ' in code_lower or 'def ' in code_lower or 'python' in code_lower:
                        language = 'python' if language == 'unknown' else language
                    elif 'function' in code_lower or 'javascript' in code_lower:
                        language = 'javascript' if language == 'unknown' else language
                    elif '{' in code_text_clean and ('"http' in code_text_clean or '"url' in code_text_clean):
                        language = 'json' if language == 'unknown' else language
                    
                    # More lenient: accept if it looks like code (has structure or is long enough)
                    looks_like_code = (
                        len(code_text_clean) > 50 or 
                        language != 'unknown' or
                        'curl' in code_lower or
                        'http' in code_lower or
                        '{' in code_text_clean or
                        '[' in code_text_clean or
                        code_text_clean.count('\n') > 2
                    )
                    
                    if looks_like_code and code_text_clean not in [ex['code'] for ex in example_code]:
                        example_code.append({
                            'language': language,
                            'code': code_text_clean[:2000]  # Limit code length
                        })
            
            # Method 2: Look for text blocks that look like code (curl commands, JSON, etc.)
            # Check all text nodes and divs that might contain code
            text_elements = soup.find_all(['div', 'p', 'span'], string=lambda text: text and len(text.strip()) > 50)
            for elem in text_elements:
                parent = elem.find_parent()
                if parent:
                    # Check if parent has code-like classes
                    parent_classes = parent.get('class', [])
                    parent_class_str = ' '.join(str(c) for c in parent_classes).lower()
                    if 'code' in parent_class_str or 'example' in parent_class_str:
                        text_content = elem.get_text().strip() if hasattr(elem, 'get_text') else str(elem).strip()
                        if text_content and len(text_content) > 50:
                            code_lower = text_content.lower()
                            # Check if it looks like a curl command or API call
                            if ('curl' in code_lower and 'http' in code_lower) or \
                               (code_lower.startswith('curl') and len(text_content.split('\n')) > 1):
                                if text_content not in [ex['code'] for ex in example_code]:
                                    example_code.append({
                                        'language': 'curl',
                                        'code': text_content[:2000]
                                    })
            
            # Method 3: Look for standalone <code> blocks (not inside <pre>)
            standalone_code_blocks = soup.find_all('code')
            for code_block in standalone_code_blocks:
                # Skip if already processed in Method 1 (inside <pre>)
                if code_block.find_parent('pre'):
                    continue
                
                code_text = code_block.get_text().strip()
                
                # Look for meaningful code (longer than single words/phrases)
                if len(code_text) > 30:
                    # Check if it looks like code (has structure)
                    has_structure = ('{' in code_text or '[' in code_text or 
                                   'http' in code_text.lower() or 
                                   'curl' in code_text.lower() or
                                   'import' in code_text.lower() or
                                   'def ' in code_text.lower() or
                                   code_text.count('\n') > 1)
                    
                    if has_structure and code_text not in [ex['code'] for ex in example_code]:
                        language = 'unknown'
                        code_lower = code_text.lower()
                        
                        # Detect language
                        if 'curl' in code_lower or code_lower.startswith('curl'):
                            language = 'curl'
                        elif 'import ' in code_lower or 'def ' in code_lower or 'python' in code_lower:
                            language = 'python'
                        elif 'function' in code_lower or 'javascript' in code_lower:
                            language = 'javascript'
                        elif '{' in code_text and ('"http' in code_text or '"url' in code_text):
                            language = 'json'
                        
                        example_code.append({
                            'language': language,
                            'code': code_text[:2000]
                        })
            
            # Method 4: Look for code blocks in specific containers
            code_containers = soup.find_all(['div', 'section'], class_=lambda x: x and ('code' in str(x).lower() or 'example' in str(x).lower()))
            for container in code_containers:
                # Get all text from container
                container_text = container.get_text().strip()
                if len(container_text) > 50 and container_text not in [ex['code'] for ex in example_code]:
                    code_lower = container_text.lower()
                    if 'curl' in code_lower or 'http' in code_lower or '{' in container_text:
                        language = 'unknown'
                        if 'curl' in code_lower:
                            language = 'curl'
                        elif 'import' in code_lower or 'python' in code_lower:
                            language = 'python'
                        elif 'function' in code_lower:
                            language = 'javascript'
                        elif '{' in container_text:
                            language = 'json'
                        
                        example_code.append({
                            'language': language,
                            'code': container_text[:2000]
                        })
                
                # Also check nested code blocks
                code_blocks = container.find_all(['pre', 'code'])
                for code_block in code_blocks:
                    code_text = code_block.get_text().strip()
                    if len(code_text) > 30 and code_text not in [ex['code'] for ex in example_code]:
                        language = 'unknown'
                        # Check for language indicators
                        if 'curl' in code_text.lower():
                            language = 'curl'
                        elif 'import' in code_text.lower() or 'python' in code_text.lower():
                            language = 'python'
                        
                        example_code.append({
                            'language': language,
                            'code': code_text[:2000]
                        })
            
            if example_code:
                endpoint_data['example_code'] = example_code
                print(f"    ✓ Extracted {len(example_code)} code examples")
            
            print(f"  ✓ Title: {endpoint_data['page_title']}")
            print(f"  ✓ Endpoint: {endpoint_data['endpoint']}")
            print(f"  ✓ Method: {endpoint_data['method']}")
            
            endpoint_data['scrape_status'] = 'success'
            
            return endpoint_data
            
        except Exception as e:
            print(f"  ✗ Error: {str(e)[:100]}")
            endpoint_data['error_message'] = str(e)[:200]
            return endpoint_data
    
    def load_existing_data(self, json_filename='tiktok_api_docs.json'):
        """Load existing scraped data from JSON file"""
        import os
        if os.path.exists(json_filename):
            try:
                with open(json_filename, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
                print(f"  ✓ Loaded {len(existing_data)} existing endpoints from {json_filename}")
                return existing_data
            except Exception as e:
                print(f"  ⚠ Could not load existing data: {e}")
                return []
        return []
    
    def calculate_endpoint_hash(self, endpoint_data):
        """Calculate a hash of endpoint content to detect changes"""
        import hashlib
        # Create a string representation of the endpoint's key data
        # Include field names, data types, and descriptions to detect all changes
        def serialize_fields(fields):
            """Serialize field list with all details"""
            return sorted([
                {
                    'field': f.get('field', ''),
                    'data_type': f.get('data_type', ''),
                    'description': f.get('description', '')[:100]  # Limit description length
                }
                for f in fields
            ], key=lambda x: x['field'])
        
        key_data = {
            'endpoint': endpoint_data.get('endpoint', ''),
            'method': endpoint_data.get('method', ''),
            'page_title': endpoint_data.get('page_title', ''),
            'headers': serialize_fields(endpoint_data.get('headers', [])),
            'parameters': serialize_fields(endpoint_data.get('parameters', [])),
            'response': serialize_fields(endpoint_data.get('response', []))
        }
        # Convert to JSON string and hash it
        key_str = json.dumps(key_data, sort_keys=True)
        return hashlib.md5(key_str.encode()).hexdigest()
    
    def find_endpoints_to_update(self, endpoint_links, existing_data):
        """
        Compare endpoint links with existing data to find:
        1. Find new endpoints (not in existing data)
        2. Find changed endpoints (content hash changed)
        3. Find unchanged endpoints (keep as-is)
        """
        from datetime import datetime
        
        # Create lookup dictionaries
        existing_by_doc_url = {item.get('doc_url', ''): item for item in existing_data}
        existing_by_endpoint = {item.get('endpoint', ''): item for item in existing_data if item.get('endpoint')}
        
        new_links = []
        changed_links = []
        unchanged_data = []
        
        for link_info in endpoint_links:
            doc_url = link_info.get('url', '')
            existing_item = existing_by_doc_url.get(doc_url)
            
            if not existing_item:
                # New endpoint - not in existing data
                new_links.append(link_info)
            else:
                # Existing endpoint - check if it changed
                # We'll need to scrape it to compare, but first check by endpoint URL
                endpoint_url = existing_item.get('endpoint', '')
                if endpoint_url:
                    # Check if endpoint URL matches (same API endpoint)
                    existing_hash = existing_item.get('content_hash')
                    if existing_hash:
                        # We'll scrape and compare hash later
                        changed_links.append(link_info)
                    else:
                        # No hash stored, assume it might have changed
                        changed_links.append(link_info)
                else:
                    # No endpoint URL stored, re-scrape to get it
                    changed_links.append(link_info)
        
        # Keep unchanged endpoints (those not in the current list)
        current_doc_urls = {link.get('url', '') for link in endpoint_links}
        for item in existing_data:
            if item.get('doc_url', '') not in current_doc_urls:
                # Endpoint removed from current list, but keep it
                unchanged_data.append(item)
            elif item.get('doc_url', '') in current_doc_urls:
                # Check if we're updating it
                link = next((l for l in endpoint_links if l.get('url') == item.get('doc_url')), None)
                if link and link not in changed_links:
                    # Not in changed list, so it's unchanged
                    unchanged_data.append(item)
        
        return new_links, changed_links, unchanged_data
    
    def update_endpoint_with_version(self, new_data, existing_item):
        """Add version and timestamp info to updated endpoint"""
        from datetime import datetime
        
        # Add version info
        version = existing_item.get('version', 1) + 1
        new_data['version'] = version
        new_data['last_updated'] = datetime.now().isoformat()
        new_data['first_scraped'] = existing_item.get('first_scraped', new_data['last_updated'])
        
        return new_data
    
    def scrape_all_endpoints(self, start_url, delay=2, save_interval=10, incremental=True, json_filename='tiktok_api_docs.json', excel_filename='tiktok_api_docs.xlsx', first_run=False):
        """
        Main scraping function with smart updates: detects changes, adds new endpoints, updates changed ones
        
        Args:
            start_url: URL to start scraping from
            delay: Delay between requests in seconds
            save_interval: Save progress every N endpoints
            incremental: If True, smart update mode (detect changes, add new)
            json_filename: JSON file to load/save existing data
            first_run: If True, override all data (first time scrape)
        """
        from datetime import datetime
        
        self.setup_driver()
        
        try:
            # Load existing data if incremental mode and not first run
            existing_data = []
            unchanged_data = []
            
            if incremental and not first_run:
                print("\n" + "="*80)
                print("STEP 0: Loading existing data")
                print("="*80)
                existing_data = self.load_existing_data(json_filename)
            
            # Get endpoint links
            endpoint_links = self.get_endpoint_links_from_page(start_url)
            
            if not endpoint_links:
                print("\n✗ No endpoint links found!")
                return existing_data if existing_data else []
            
            # Smart update: find new, changed, and unchanged endpoints
            if incremental and not first_run and existing_data:
                print("\n" + "="*80)
                print("STEP 1.5: Analyzing endpoints (detecting changes)")
                print("="*80)
                
                new_links, changed_links, unchanged_data = self.find_endpoints_to_update(
                    endpoint_links, existing_data
                )
                
                print(f"  ✓ New endpoints to add: {len(new_links)}")
                print(f"  ✓ Changed endpoints to update: {len(changed_links)}")
                print(f"  ✓ Unchanged endpoints: {len(unchanged_data)}")
                
                # Combine new and changed links to scrape
                links_to_scrape = new_links + changed_links
                
                if not links_to_scrape:
                    print("\n✓ All endpoints are up to date! No changes detected.")
                    return unchanged_data
                
                endpoint_links = links_to_scrape
            else:
                print(f"\n  → Scraping all {len(endpoint_links)} endpoints")
                if first_run:
                    print("  → First run mode: Will override all existing data")
            
            print("\n" + "="*80)
            print("STEP 2: Scraping endpoint details")
            print("="*80 + "\n")
            
            updated_data = []
            failed_count = 0
            success_count = 0
            new_count = 0
            changed_count = 0
            
            # Create lookup for existing items (for versioning)
            existing_by_doc_url = {item.get('doc_url', ''): item for item in existing_data}
            
            for i, link_info in enumerate(endpoint_links, 1):
                doc_url = link_info.get('url', '')
                existing_item = existing_by_doc_url.get(doc_url)
                is_new = existing_item is None
                
                print(f"[{i}/{len(endpoint_links)}] {link_info['text']}")
                if is_new:
                    print("  📌 NEW endpoint")
                else:
                    print("  🔄 UPDATING existing endpoint")
                
                endpoint_data = self.extract_endpoint_details(link_info['url'], link_info['text'])
                
                # Add versioning and timestamp info
                if endpoint_data['scrape_status'] == 'success':
                    if is_new:
                        # New endpoint
                        endpoint_data['version'] = 1
                        endpoint_data['first_scraped'] = datetime.now().isoformat()
                        endpoint_data['last_updated'] = datetime.now().isoformat()
                        new_count += 1
                    else:
                        # Updated endpoint - check if actually changed
                        old_hash = existing_item.get('content_hash', '')
                        new_hash = self.calculate_endpoint_hash(endpoint_data)
                        endpoint_data['content_hash'] = new_hash
                        
                        if old_hash and old_hash == new_hash:
                            # No actual change detected - keep existing data
                            print(f"    ✓ No changes detected (hash match)")
                            endpoint_data = existing_item.copy()  # Keep old data
                            endpoint_data['last_checked'] = datetime.now().isoformat()
                            # Don't increment changed_count - it's unchanged
                        else:
                            # Actually changed - show what changed
                            print(f"    ✓ Changes detected!")
                            if not old_hash:
                                print(f"      → First time calculating hash for this endpoint")
                            else:
                                # Compare to show what changed
                                old_fields = {
                                    'headers': len(existing_item.get('headers', [])),
                                    'parameters': len(existing_item.get('parameters', [])),
                                    'response': len(existing_item.get('response', []))
                                }
                                new_fields = {
                                    'headers': len(endpoint_data.get('headers', [])),
                                    'parameters': len(endpoint_data.get('parameters', [])),
                                    'response': len(endpoint_data.get('response', []))
                                }
                                changes = []
                                for field_type in ['headers', 'parameters', 'response']:
                                    if old_fields[field_type] != new_fields[field_type]:
                                        changes.append(f"{field_type}: {old_fields[field_type]} → {new_fields[field_type]}")
                                if changes:
                                    print(f"      → Field counts changed: {', '.join(changes)}")
                                else:
                                    print(f"      → Field details changed (types/descriptions)")
                            
                            # Actually changed - increment version
                            endpoint_data = self.update_endpoint_with_version(endpoint_data, existing_item)
                            endpoint_data['content_hash'] = new_hash
                            changed_count += 1
                    
                    success_count += 1
                else:
                    failed_count += 1
                    print(f"  ⚠ Status: FAILED")
                    # Keep existing data if available
                    if existing_item:
                        endpoint_data = existing_item
                
                updated_data.append(endpoint_data)
                print()
                
                # Save progress periodically to the FINAL files (overwrite each time)
                if i % save_interval == 0:
                    combined_data = unchanged_data + updated_data
                    print(f"  💾 Saving progress to final files... ({i}/{len(endpoint_links)} endpoints)")
                    
                    # Save to final JSON file (overwrite)
                    try:
                        self.save_to_json(combined_data, json_filename)
                    except Exception as e:
                        print(f"  ⚠ JSON save failed: {e}")
                    
                    # Save API config JSON (for making API calls) - overwrite
                    try:
                        config_filename = 'tiktok_api_config.json'
                        self.save_api_config_json(combined_data, config_filename)
                    except Exception as e:
                        print(f"  ⚠ API config JSON save failed: {e}")
                    
                    # Save to final Excel file (overwrite)
                    try:
                        self.save_to_excel(combined_data, excel_filename, mode='overwrite')
                        print(f"  ✓ Progress saved to {json_filename}, {config_filename}, and {excel_filename}")
                    except Exception as e:
                        print(f"  ⚠ Excel save failed: {e}, but JSON saved")
                    
                    print()
                
                # Add delay between requests
                if i < len(endpoint_links):
                    time.sleep(delay)
            
            # Combine unchanged + updated data
            all_data = unchanged_data + updated_data
            
            # Sort by endpoint name for consistency
            all_data.sort(key=lambda x: x.get('endpoint_name', ''))
            
            # Print final statistics
            print("\n" + "="*80)
            print("SCRAPING COMPLETE")
            print("="*80)
            if incremental and not first_run:
                print(f"✓ Unchanged endpoints: {len(unchanged_data)}")
            print(f"✓ New endpoints added: {new_count}")
            print(f"✓ Updated endpoints: {changed_count}")
            print(f"✓ Successfully scraped: {success_count}/{len(endpoint_links)}")
            if failed_count > 0:
                print(f"⚠ Failed to scrape: {failed_count}/{len(endpoint_links)}")
            print(f"✓ Total endpoints: {len(all_data)}")
            print(f"📅 Last update: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print()
            
            return all_data
            
        finally:
            if self.driver:
                self.driver.quit()
    
    def save_to_json(self, data, filename='tiktok_api_docs.json'):
        """Save to JSON"""
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"\n✓ JSON data saved to {filename}")
    
    def parse_curl_example(self, curl_code: str):
        """
        Parse curl example code to extract complete call structure
        Returns dict with complete curl structure for dynamic execution
        """
        import re
        import json as json_module
        
        result = {
            'method': 'GET',
            'url_template': '',  # URL with placeholders
            'url_base': '',  # Base URL without query params
            'headers': {},  # Headers with placeholders
            'body_template': None,  # Body template with placeholders
            'query_params_template': {},  # Query params with placeholders
            'body_format': 'json',  # json, form-data, raw
            'has_query_params': False,
            'has_body': False
        }
        
        curl_lower = curl_code.lower()
        
        # Extract method
        method_match = re.search(r"--request\s+(\w+)", curl_code, re.IGNORECASE)
        if method_match:
            result['method'] = method_match.group(1).upper()
        
        # Extract URL (may contain query params)
        url_match = re.search(r"'(https?://[^']+)'", curl_code)
        if url_match:
            full_url = url_match.group(1)
            result['url_template'] = full_url
            
            # Split URL and query params
            if '?' in full_url:
                result['url_base'] = full_url.split('?')[0]
                query_part = full_url.split('?', 1)[1]
                result['has_query_params'] = True
                
                # Parse query parameters (handle arrays, dates, etc.)
                # Example: fields=["item_id","create_time"]&start_date=2021-07-28
                query_pairs = query_part.split('&')
                for pair in query_pairs:
                    if '=' in pair:
                        key, value = pair.split('=', 1)
                        # Check if value is an array format
                        if value.startswith('[') and value.endswith(']'):
                            # Array parameter: fields=["item_id","create_time"]
                            result['query_params_template'][key] = {
                                'type': 'array',
                                'template': value  # Keep template with placeholders
                            }
                        else:
                            # Regular parameter
                            result['query_params_template'][key] = {
                                'type': 'string',
                                'template': value  # May contain {{placeholder}}
                            }
            else:
                result['url_base'] = full_url
        
        # Extract headers
        header_matches = re.findall(r"--header\s+'([^:]+):\s*([^']+)'", curl_code)
        for header_name, header_value in header_matches:
            result['headers'][header_name.strip()] = header_value.strip()
        
        # Extract body (multiple formats possible)
        # --data-raw
        data_raw_match = re.search(r"--data-raw\s+'([^']+)'", curl_code, re.DOTALL)
        if data_raw_match:
            body_content = data_raw_match.group(1)
            result['has_body'] = True
            result['body_template'] = body_content
            result['body_format'] = 'json'
        
        # --data
        data_match = re.search(r"--data\s+'([^']+)'", curl_code, re.DOTALL)
        if data_match and not result['has_body']:
            body_content = data_match.group(1)
            result['has_body'] = True
            result['body_template'] = body_content
            result['body_format'] = 'form-data'
        
        # --form (multipart form data)
        form_matches = re.findall(r"--form\s+'([^']+)'", curl_code)
        if form_matches:
            result['has_body'] = True
            result['body_format'] = 'multipart'
            result['body_template'] = form_matches
        
        return result
    
    def generate_api_id(self, endpoint_url: str, api_name: str, page_title: str) -> str:
        """
        Generate a unique API ID from endpoint URL or API name
        Examples:
        - /open_api/v1.3/business/get/ -> business_get
        - /open_api/v1.3/business/comment/create/ -> business_comment_create
        - /open_api/v1.3/tt_user/token_info/get/ -> tt_user_token_info_get
        """
        import re
        
        # Try to extract from endpoint URL first
        if endpoint_url:
            # Extract path after /open_api/
            match = re.search(r'/open_api/[v\d\.]+/([^/?]+)', endpoint_url)
            if match:
                path_parts = match.group(1).strip('/').split('/')
                # Convert to snake_case
                api_id = '_'.join(path_parts)
                # Clean up: remove version numbers, special chars
                api_id = re.sub(r'[^a-zA-Z0-9_]', '_', api_id)
                api_id = re.sub(r'_+', '_', api_id).strip('_')
                if api_id:
                    return api_id
        
        # Fallback: generate from API name or page title
        source = page_title or api_name or 'unknown'
        # Convert to lowercase, replace spaces/special chars with underscores
        api_id = re.sub(r'[^a-zA-Z0-9\s]', '', source.lower())
        api_id = re.sub(r'\s+', '_', api_id)
        api_id = re.sub(r'_+', '_', api_id).strip('_')
        
        # Limit length
        if len(api_id) > 50:
            api_id = api_id[:50]
        
        return api_id if api_id else 'unknown_api'
    
    def save_api_config_json(self, data, filename='tiktok_api_config.json'):
        """
        Save API call configuration JSON - ready to use for making API calls
        Each endpoint contains all info needed: method, URL, headers, parameters
        Parses example code to determine call format dynamically
        """
        config_data = []
        
        for item in data:
            if item.get('scrape_status') != 'success':
                continue  # Skip failed endpoints
            
            endpoint_url = item.get('endpoint', '')
            method = item.get('method', 'GET')
            page_title = item.get('page_title', item.get('endpoint_name', ''))
            
            if not endpoint_url or endpoint_url == 'NOT FOUND':
                continue  # Skip endpoints without URL
            
            # Parse example code to get complete call structure
            example_code = item.get('example_code', [])
            curl_structure = None
            
            # Try to parse curl examples - store the complete structure
            for example in example_code:
                if example.get('language') == 'curl' and example.get('code'):
                    curl_structure = self.parse_curl_example(example['code'])
                    # Use parsed method if available
                    if curl_structure.get('method'):
                        method = curl_structure['method']
                    break
            
            # If no curl example found, create default structure
            if not curl_structure:
                curl_structure = {
                    'method': method.upper(),
                    'url_template': endpoint_url,
                    'url_base': endpoint_url,
                    'headers': {},
                    'body_template': None,
                    'query_params_template': {},
                    'body_format': 'json',
                    'has_query_params': False,
                    'has_body': False
                }
            
            # Build headers dict with placeholders
            headers = {}
            for header in item.get('headers', []):
                field_name = header.get('field', '')
                if field_name:
                    # Use placeholder for Access-Token, actual value for others
                    if 'token' in field_name.lower() or 'access' in field_name.lower():
                        headers[field_name] = '{{Access-Token}}'
                    else:
                        # Get default value if available, otherwise use placeholder
                        default_value = header.get('description', '')
                        if default_value and default_value.lower() not in ['required', 'optional']:
                            headers[field_name] = default_value
                        else:
                            headers[field_name] = f'{{{{{field_name}}}}}'
            
            # Build parameters dict with placeholders and type info
            parameters = {}
            parameter_types = {}  # Track parameter types for array handling
            parameter_supported_fields = {}  # Track supported fields for array parameters
            for param in item.get('parameters', []):
                field_name = param.get('field', '')
                if field_name:
                    data_type = param.get('data_type', '').lower()
                    # Use placeholder
                    parameters[field_name] = f'{{{{{field_name}}}}}'
                    
                    # Track parameter types
                    if 'array' in data_type or '[]' in data_type:
                        parameter_types[field_name] = 'array'
                    elif 'date' in field_name.lower():
                        parameter_types[field_name] = 'date'
                    elif 'fields' in field_name.lower():
                        parameter_types[field_name] = 'array'  # fields is usually an array
                    
                    # Extract supported fields if available (for array parameters like "fields")
                    supported_fields = param.get('supported_fields', [])
                    if supported_fields:
                        parameter_supported_fields[field_name] = supported_fields
            
            # Generate API ID from endpoint URL or API name
            api_id = self.generate_api_id(endpoint_url, item.get('endpoint_name', ''), page_title)
            
            # Generate API ID from endpoint URL or API name
            api_id = self.generate_api_id(endpoint_url, item.get('endpoint_name', ''), page_title)
            
            config_item = {
                'api_id': api_id,  # Unique identifier for API selection
                'api_name': item.get('endpoint_name', ''),
                'page_title': page_title,
                'method': curl_structure['method'],
                'endpoint_url': endpoint_url,
                'headers': headers,
                'parameters': parameters,
                'parameter_types': parameter_types,  # Add type info
                'parameter_supported_fields': parameter_supported_fields,  # Supported field names for array params
                'curl_structure': curl_structure,  # Complete curl structure for dynamic execution
                'response_fields': [r.get('field', '') for r in item.get('response', [])],
                'doc_url': item.get('doc_url', ''),
                'example_code': example_code
            }
            
            config_data.append(config_item)
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, indent=2, ensure_ascii=False)
        print(f"\n✓ API Config JSON saved to {filename} ({len(config_data)} APIs)")
    
    def save_to_excel(self, data, filename='tiktok_api_docs.xlsx', mode='overwrite', update_existing=False):
        """
        Save to Excel with unified metrics, matrix sheets, detailed fields, and failure marking
        
        Args:
            data: List of endpoint data dictionaries
            filename: Output Excel filename
            mode: 'overwrite' (default) - replaces entire file
                  'append' - appends new data to existing sheets (if file exists)
                  'backup' - creates backup before overwriting
        """
        import os
        from datetime import datetime
        
        # Handle backup mode
        if mode == 'backup' and os.path.exists(filename):
            backup_name = f"{filename.rsplit('.', 1)[0]}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            import shutil
            shutil.copy2(filename, backup_name)
            print(f"  💾 Created backup: {backup_name}")
        
        # Handle append mode - load existing workbook
        if mode == 'append' and os.path.exists(filename):
            from openpyxl import load_workbook
            print(f"  📝 Appending to existing file: {filename}")
            workbook = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace')
            writer.book = workbook
        else:
            # Overwrite mode (default)
            print("Creating Excel file with multiple sheets...")
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='w')
        
        # Create summary with status marking
        summary_data = []
        for item in data:
            summary_data.append({
                'Status': item.get('scrape_status', 'unknown').upper(),
                'Endpoint Name': item.get('endpoint_name', ''),
                'Page Title': item.get('page_title', ''),
                'Documentation URL': item.get('doc_url', ''),
                'API Endpoint': item.get('endpoint', 'NOT FOUND'),
                'Method': item.get('method', 'NOT FOUND'),
                'Headers Count': len(item.get('headers', [])),
                'Parameters Count': len(item.get('parameters', [])),
                'Response Fields Count': len(item.get('response', [])),
                'Error Message': item.get('error_message', '')
            })
        
        with writer:
            # 1. SUMMARY SHEET
            print("  Creating Summary sheet...")
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Apply formatting to highlight failed rows
            workbook = writer.book
            worksheet = writer.sheets['Summary']
            
            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            
            for row_idx, row in enumerate(summary_df.itertuples(), start=2):
                if row.Status == 'FAILED':
                    for col_idx in range(1, len(summary_df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
                elif row.Status == 'SUCCESS':
                    worksheet.cell(row=row_idx, column=1).fill = green_fill
            
            # 2. API FIELDS DETAIL SHEET - Shows actual field names for each API
            print("  Creating API Fields Detail sheet...")
            self._create_detailed_fields_sheet(data, writer)
            
            # 3. MATRIX SHEETS - Rows = API URLs, Columns = Metrics, X marks
            print("  Creating Matrix sheets...")
            successful_data = [item for item in data if item.get('scrape_status') == 'success']
            if successful_data:
                self._create_matrix_sheets(successful_data, writer)
            
            # 4. UNIFIED METRICS SHEETS (only for successful scrapes)
            print("  Creating Unified Metrics sheets...")
            if successful_data:
                self._create_unified_metrics_sheets(successful_data, writer)
            
            # 5. INDIVIDUAL ENDPOINT SHEETS
            print("  Creating individual endpoint sheets...")
            for i, item in enumerate(data, 1):
                # Sanitize sheet name - Excel doesn't allow: / \ ? * [ ]
                sheet_name = item.get('endpoint_name', f'Endpoint_{i}')
                # Replace invalid characters
                sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_')
                sheet_name = sheet_name.replace('*', '_').replace('[', '_').replace(']', '_')
                sheet_name = sheet_name.replace(':', '_').replace(';', '_')
                # Limit to 31 characters (Excel limit)
                sheet_name = sheet_name[:31]
                
                details_data = []
                details_data.append(['Status', item.get('scrape_status', 'unknown').upper()])
                details_data.append(['Endpoint Name', item.get('endpoint_name', '')])
                details_data.append(['Page Title', item.get('page_title', '')])
                details_data.append(['Documentation URL', item.get('doc_url', '')])
                details_data.append(['API Endpoint', item.get('endpoint', 'NOT FOUND')])
                details_data.append(['Method', item.get('method', 'NOT FOUND')])
                
                if item.get('error_message'):
                    details_data.append(['Error Message', item.get('error_message', '')])
                
                details_data.append(['', ''])
                
                # Example code section
                if item.get('example_code'):
                    details_data.append(['EXAMPLE CODE', '', ''])
                    for idx, example in enumerate(item['example_code'], 1):
                        details_data.append([f'Example {idx} ({example.get("language", "unknown")})', '', ''])
                        # Split long code into multiple rows for readability
                        code_lines = example.get('code', '').split('\n')
                        for line in code_lines[:50]:  # Limit to 50 lines per example
                            details_data.append(['', '', line[:200]])  # Limit line length
                        details_data.append(['', '', ''])
                    details_data.append(['', '', ''])
                
                if item.get('headers'):
                    details_data.append(['HEADERS', '', '', ''])
                    details_data.append(['Field', 'Data Type', 'Description', ''])
                    for h in item['headers']:
                        details_data.append([
                            h.get('field', ''),
                            h.get('data_type', ''),
                            h.get('description', ''),
                            ''  # Empty column for consistency
                        ])
                    details_data.append(['', '', '', ''])
                else:
                    details_data.append(['HEADERS', '', '', ''])
                    details_data.append(['No headers found in documentation', '', '', ''])
                    details_data.append(['', '', '', ''])
                
                if item.get('parameters'):
                    details_data.append(['PARAMETERS', '', '', ''])
                    details_data.append(['Field', 'Data Type', 'Description', 'Supported Fields'])
                    for p in item['parameters']:
                        field_name = p.get('field', '')
                        data_type = p.get('data_type', '')
                        description = p.get('description', '')
                        supported_fields = p.get('supported_fields', [])
                        
                        if supported_fields:
                            # Show parameter with supported fields in the new column
                            supported_fields_str = ', '.join(supported_fields)
                            details_data.append([
                                field_name,
                                data_type,
                                description,
                                supported_fields_str
                            ])
                        else:
                            # Regular parameter without supported fields
                            details_data.append([
                                field_name,
                                data_type,
                                description,
                                ''  # Empty supported fields column
                            ])
                    details_data.append(['', '', '', ''])
                else:
                    details_data.append(['PARAMETERS', '', '', ''])
                    details_data.append(['No parameters found in documentation', '', '', ''])
                    details_data.append(['', '', '', ''])
                
                if item.get('response'):
                    details_data.append(['RESPONSE', '', ''])
                    details_data.append(['Field', 'Data Type', 'Description'])
                    for r in item['response']:
                        details_data.append([r.get('field', ''), r.get('data_type', ''), r.get('description', '')])
                
                df = pd.DataFrame(details_data)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        print(f"✓ Excel data saved to {filename}")
        
        # Print sheet summary
        print("\n" + "="*80)
        print("EXCEL SHEETS CREATED")
        print("="*80)
        print("1. Summary - Overview with counts")
        print("2. API Fields Detail - Field names for each API")
        print("3. Headers Matrix - Which APIs have which headers (X marks)")
        print("4. Parameters Matrix - Which APIs have which parameters (X marks)")
        print("5. Response Matrix - Which APIs have which response fields (X marks)")
        print("6. All Headers - Unified headers analysis")
        print("7. All Parameters - Unified parameters analysis")
        print("8. All Response Fields - Unified response fields analysis")
        print("9+. Individual endpoint sheets (one per API)")
        
        # Print failure summary
        failed_endpoints = [item for item in data if item.get('scrape_status') == 'failed']
        if failed_endpoints:
            print(f"\n⚠ Warning: {len(failed_endpoints)} endpoints failed to scrape:")
            for item in failed_endpoints[:10]:  # Show first 10
                print(f"  - {item['endpoint_name']}")
            if len(failed_endpoints) > 10:
                print(f"  ... and {len(failed_endpoints) - 10} more (see Summary sheet)")
            print(f"\n  Failed endpoints are highlighted in RED in the Summary sheet")
    
    def _create_detailed_fields_sheet(self, data, writer):
        """Create sheet with actual field names for each API"""
        rows = []
        
        for item in data:
            header_names = ', '.join([h.get('field', '') for h in item.get('headers', [])])
            
            # For parameters, expand supported fields
            param_list = []
            for p in item.get('parameters', []):
                field_name = p.get('field', '')
                supported_fields = p.get('supported_fields', [])
                if supported_fields:
                    # Show parameter name and its supported fields
                    param_list.append(f"{field_name} ({', '.join(supported_fields)})")
                else:
                    param_list.append(field_name)
            param_names = ', '.join(param_list)
            
            response_names = ', '.join([r.get('field', '') for r in item.get('response', [])])
            
            rows.append({
                'Status': item.get('scrape_status', 'unknown').upper(),
                'Endpoint Name': item.get('endpoint_name', ''),
                'Page Title': item.get('page_title', ''),
                'API Endpoint': item.get('endpoint', 'NOT FOUND'),
                'Method': item.get('method', 'NOT FOUND'),
                'Header Fields': header_names if header_names else 'None',
                'Parameter Fields': param_names if param_names else 'None',
                'Response Fields': response_names if response_names else 'None',
                'Documentation URL': item.get('doc_url', '')
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='API Fields Detail', index=False)
        
        # Apply color coding
        worksheet = writer.sheets['API Fields Detail']
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        
        for row_idx, row in enumerate(df.itertuples(), start=2):
            if row.Status == 'FAILED':
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
            elif row.Status == 'SUCCESS':
                worksheet.cell(row=row_idx, column=1).fill = green_fill
        
        print(f"    ✓ API Fields Detail: {len(df)} APIs")
    
    def _create_matrix_sheets(self, data, writer):
        """Create matrix sheets for headers, parameters, and response fields"""
        if not data:
            print("    ⚠ No successful data for matrix")
            return
        
        self._create_single_matrix(data, writer, 'headers', 'Headers Matrix')
        self._create_single_matrix(data, writer, 'parameters', 'Parameters Matrix')
        self._create_single_matrix(data, writer, 'response', 'Response Matrix')
    
    def _create_single_matrix(self, data, writer, field_type, sheet_name):
        """Create a single matrix sheet"""
        # Collect all unique fields
        # For parameters, expand supported fields
        all_fields = set()
        for item in data:
            for field_data in item.get(field_type, []):
                field_name = field_data.get('field', '')
                if field_name:
                    if field_type == 'parameters':
                        # For parameters, check if there are supported fields
                        supported_fields = field_data.get('supported_fields', [])
                        if supported_fields:
                            # Add each supported field as a separate column
                            for sf in supported_fields:
                                all_fields.add(sf)
                        else:
                            # Regular parameter without supported fields
                            all_fields.add(field_name)
                    else:
                        # Headers and response fields: use field name as-is
                        all_fields.add(field_name)
        
        sorted_fields = sorted(all_fields)
        
        if not sorted_fields:
            print(f"    ⚠ No fields for {sheet_name}")
            return
        
        # Create matrix
        matrix_rows = []
        
        for item in data:
            api_name = item.get('endpoint_name', '')
            api_url = item.get('endpoint', 'NOT FOUND')
            method = item.get('method', '')
            
            # Get fields this API has
            if field_type == 'parameters':
                # For parameters, expand supported fields
                api_fields = set()
                for field_data in item.get(field_type, []):
                    field_name = field_data.get('field', '')
                    supported_fields = field_data.get('supported_fields', [])
                    if supported_fields:
                        # Add each supported field
                        api_fields.update(supported_fields)
                    else:
                        # Regular parameter
                        api_fields.add(field_name)
            else:
                # Headers and response: use field names as-is
                api_fields = set([f.get('field', '') for f in item.get(field_type, [])])
            
            # Create row
            row_data = {
                'API Endpoint Name': api_name,
                'API URL': api_url,
                'Method': method
            }
            
            # Add X for each field
            for field in sorted_fields:
                row_data[field] = 'X' if field in api_fields else ''
            
            matrix_rows.append(row_data)
        
        # Write to Excel
        df = pd.DataFrame(matrix_rows)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Format
        worksheet = writer.sheets[sheet_name]
        from openpyxl.styles import Alignment, Font, PatternFill
        
        center_aligned = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Format header row
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
        
        # Center align and color cells with X marks
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White for empty
        
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(4, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = center_aligned
                if cell.value == 'X':
                    cell.font = Font(bold=True, color='006100')  # Dark green text
                    cell.fill = green_fill  # Green background for entire cell
                else:
                    # White background for empty cells
                    cell.fill = white_fill
        
        # Auto-adjust column widths
        for col_idx in range(1, 4):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 30
        for col_idx in range(4, len(df.columns) + 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 15
        
        print(f"    ✓ {sheet_name}: {len(df)} APIs × {len(sorted_fields)} fields")
    
    def _create_unified_metrics_sheets(self, data, writer):
        """Create unified metrics sheets"""
        # [Same implementation as in the requests-based scraper]
        # Collect unique fields
        headers_map = {}
        parameters_map = {}
        response_map = {}
        
        for item in data:
            endpoint_info = {
                'endpoint_name': item.get('endpoint_name', ''),
                'api_endpoint': item.get('endpoint', ''),
                'doc_url': item.get('doc_url', ''),
                'method': item.get('method', '')
            }
            
            for header in item.get('headers', []):
                field = header.get('field', '')
                if field:
                    if field not in headers_map:
                        headers_map[field] = {
                            'data_type': header.get('data_type', ''),
                            'description': header.get('description', ''),
                            'endpoints': []
                        }
                    headers_map[field]['endpoints'].append(endpoint_info)
            
            for param in item.get('parameters', []):
                field = param.get('field', '')
                supported_fields = param.get('supported_fields', [])
                
                if supported_fields:
                    # For parameters with supported fields, add each supported field
                    for sf in supported_fields:
                        if sf not in parameters_map:
                            parameters_map[sf] = {
                                'data_type': param.get('data_type', ''),
                                'description': f"Supported field for {field}: {param.get('description', '')}",
                                'endpoints': []
                            }
                        parameters_map[sf]['endpoints'].append(endpoint_info)
                elif field:
                    # Regular parameter without supported fields
                    if field not in parameters_map:
                        parameters_map[field] = {
                            'data_type': param.get('data_type', ''),
                            'description': param.get('description', ''),
                            'endpoints': []
                        }
                    parameters_map[field]['endpoints'].append(endpoint_info)
            
            for resp in item.get('response', []):
                field = resp.get('field', '')
                if field:
                    if field not in response_map:
                        response_map[field] = {
                            'data_type': resp.get('data_type', ''),
                            'description': resp.get('description', ''),
                            'endpoints': []
                        }
                    response_map[field]['endpoints'].append(endpoint_info)
        
        # Create sheets (same format as before)
        for field_map, sheet_name in [
            (headers_map, 'All Headers'),
            (parameters_map, 'All Parameters'),
            (response_map, 'All Response Fields')
        ]:
            if field_map:
                rows = []
                for field, info in sorted(field_map.items()):
                    for idx, ep in enumerate(info['endpoints']):
                        rows.append({
                            'Field': field if idx == 0 else '',
                            'Data Type': info['data_type'] if idx == 0 else '',
                            'Description': info['description'] if idx == 0 else '',
                            'Usage Count': len(info['endpoints']) if idx == 0 else '',
                            'Endpoint Name': ep['endpoint_name'],
                            'API Endpoint URL': ep['api_endpoint'],
                            'Method': ep['method'],
                            'Documentation URL': ep['doc_url']
                        })
                    rows.append({k: '' for k in rows[0].keys()})  # Empty row
                
                pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"  ✓ Created unified metrics sheets")


def main():
    print("\n" + "="*80)
    print("TikTok Business API Documentation Scraper (Selenium)")
    print("="*80 + "\n")
    
    scraper = TikTokAPIDocScraperSelenium(headless=True)
    start_url = "https://business-api.tiktok.com/portal/docs?id=1735713875563521"
    
    # First run: Set first_run=True to override all data
    # Subsequent runs: Set first_run=False for smart updates
    first_run = False  # Change to True for first time scrape
    
    data = scraper.scrape_all_endpoints(
        start_url, 
        delay=2, 
        save_interval=10,  # Save both JSON and Excel every 10 endpoints
        incremental=True,  # Smart update mode
        json_filename='tiktok_api_docs.json',
        excel_filename='tiktok_api_docs.xlsx',
        first_run=first_run  # True = override all, False = smart update
    )
    
    if data:
        # Calculate statistics
        total = len(data)
        successful = len([d for d in data if d.get('scrape_status') == 'success'])
        failed = len([d for d in data if d.get('scrape_status') == 'failed'])
        
        print("\n" + "="*80)
        print(f"RESULTS: Scraped {total} endpoints")
        print("="*80)
        print(f"✓ Successful: {successful}")
        if failed > 0:
            print(f"⚠ Failed: {failed}")
        print()
        
        print("Saving data...")
        
        # Save final JSON (always updates with merged data)
        scraper.save_to_json(data)
        
        # Save API config JSON (for making API calls)
        scraper.save_api_config_json(data, 'tiktok_api_config.json')
        
        # Save final Excel (with versioning info)
        # Note: Progress Excel files are already saved during scraping
        try:
            scraper.save_to_excel(data, mode='overwrite', update_existing=True)
            print("✓ Final Excel file saved")
        except Exception as e:
            print(f"⚠ Final Excel save failed: {e}")
            print("  → But progress Excel files are available from batch saves")
        
        print("\n" + "="*80)
        print("OUTPUT FILES")
        print("="*80)
        print("✓ tiktok_api_docs.json - Complete data (includes failed attempts)")
        print("✓ tiktok_api_config.json - API call config (ready to use for API calls)")
        print("✓ tiktok_api_docs.xlsx - Excel with multiple sheets:")
        print("  - Summary: Overview with status indicators")
        print("    • GREEN = Successfully scraped")
        print("    • RED = Failed to scrape")
        print("  - All Headers: Unified list (successful endpoints only)")
        print("  - All Parameters: Unified list (successful endpoints only)")
        print("  - All Response Fields: Unified list (successful endpoints only)")
        print("  - Individual endpoint sheets with full details")
        
        if failed > 0:
            print(f"\n⚠ Note: {failed} endpoints failed to scrape.")
            print("  Check the Summary sheet for details.")
            print("  Failed endpoints are highlighted in RED.")
            print("  You can retry these manually or run the scraper again.")
    else:
        print("\n✗ No data was scraped.")


def convert_json_to_excel(json_filename='tiktok_api_docs.json', excel_filename='tiktok_api_docs.xlsx'):
    """
    Utility function to convert existing JSON file to Excel
    Use this if scraping completed but Excel save failed
    """
    import os
    
    if not os.path.exists(json_filename):
        print(f"✗ JSON file not found: {json_filename}")
        return False
    
    print(f"Loading data from {json_filename}...")
    scraper = TikTokAPIDocScraperSelenium(headless=True)
    
    try:
        # Load JSON data
        with open(json_filename, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print(f"✓ Loaded {len(data)} endpoints from JSON")
        
        # Save to Excel
        print(f"Converting to Excel: {excel_filename}...")
        scraper.save_to_excel(data, excel_filename, mode='overwrite')
        
        print(f"\n✓ Successfully converted JSON to Excel!")
        print(f"  → {len(data)} endpoints saved to {excel_filename}")
        return True
        
    except Exception as e:
        print(f"✗ Error converting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    import sys
    
    # Check if user wants to convert JSON to Excel
    if len(sys.argv) > 1 and sys.argv[1] == '--convert':
        json_file = sys.argv[2] if len(sys.argv) > 2 else 'tiktok_api_docs.json'
        excel_file = sys.argv[3] if len(sys.argv) > 3 else 'tiktok_api_docs.xlsx'
        convert_json_to_excel(json_file, excel_file)
    else:
        main()