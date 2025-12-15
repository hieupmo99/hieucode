import json
import time
import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup

class TikTokAPIDocScraperSelenium:
    def __init__(self, headless=True):
        """
        Initialize Selenium-based scraper for JavaScript-rendered content
        """
        self.base_url = "https://business-api.tiktok.com"
        self.driver = None
        self.headless = headless
        
    def setup_driver(self):
        """
        Setup Chrome driver with appropriate options
        """
        chrome_options = Options()
        if self.headless:
            chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        # Try common ChromeDriver paths
        chromedriver_paths = [
            "/opt/homebrew/bin/chromedriver",  # Homebrew on Apple Silicon
            "/usr/local/bin/chromedriver",      # Homebrew on Intel Mac
            "chromedriver"                      # System PATH
        ]
        
        driver_service = None
        for path in chromedriver_paths:
            try:
                from selenium.webdriver.chrome.service import Service
                driver_service = Service(executable_path=path)
                self.driver = webdriver.Chrome(service=driver_service, options=chrome_options)
                print(f"✓ Using ChromeDriver at: {path}\n")
                break
            except Exception as e:
                continue
        
        # Fallback to default if none of the paths work
        if not self.driver:
            try:
                self.driver = webdriver.Chrome(options=chrome_options)
                print("✓ Using ChromeDriver from system PATH\n")
            except Exception as e:
                print(f"✗ Error: Could not initialize ChromeDriver")
                print(f"  {e}")
                print("\nPlease ensure ChromeDriver is installed:")
                print("  brew install chromedriver")
                raise
        
        self.driver.set_page_load_timeout(30)
        
    def get_iframe_content(self, url, wait_time=10, max_retries=3):
        """
        Load page and extract iframe content using Selenium with retry logic
        """
        for attempt in range(max_retries):
            try:
                print(f"Loading page: {url}")
                
                # Check if driver is still alive
                try:
                    _ = self.driver.current_url
                except Exception:
                    print("  ⚠ Driver session lost, recreating...")
                    self.driver.quit()
                    self.setup_driver()
                
                self.driver.get(url)
                
                # Wait for iframe to load
                print("Waiting for iframe to load...")
                iframe = WebDriverWait(self.driver, wait_time).until(
                    EC.presence_of_element_located((By.TAG_NAME, "iframe"))
                )
                
                print("✓ Iframe found!")
                
                # Get iframe src
                iframe_src = iframe.get_attribute('src')
                print(f"Iframe URL: {iframe_src[:100]}...")
                
                # Switch to iframe
                self.driver.switch_to.frame(iframe)
                
                # Wait for content to load
                time.sleep(2)
                
                # Get page source from iframe
                iframe_html = self.driver.page_source
                
                # Switch back to main content
                self.driver.switch_to.default_content()
                
                return iframe_html, iframe_src
                
            except Exception as e:
                print(f"  ⚠ Attempt {attempt + 1}/{max_retries} failed: {str(e)[:100]}")
                
                if attempt < max_retries - 1:
                    print(f"  → Retrying in 3 seconds...")
                    time.sleep(3)
                    
                    # Try to recover the driver
                    try:
                        self.driver.switch_to.default_content()
                    except:
                        pass
                    
                    try:
                        self.driver.quit()
                    except:
                        pass
                    
                    self.setup_driver()
                else:
                    print(f"  ✗ All retry attempts failed")
                    return None, None
        
        return None, None
    
    def get_endpoint_links_from_page(self, page_url):
        """
        Extract all endpoint links from the documentation page
        """
        print("\n" + "="*80)
        print("STEP 1: Getting list of endpoint documentation pages")
        print("="*80)
        print(f"\nFetching endpoint links from: {page_url}")
        
        iframe_html, iframe_src = self.get_iframe_content(page_url)
        
        if not iframe_html:
            print("✗ Could not load iframe content")
            return []
        
        soup = BeautifulSoup(iframe_html, 'html.parser')
        
        endpoint_links = []
        all_links = soup.find_all('a', href=True)
        
        print(f"Found {len(all_links)} total links in iframe")
        
        for link in all_links:
            href = link.get('href')
            text = link.get_text(strip=True)
            
            # Look for documentation links with id parameter
            if href and 'id=' in href and text and len(text) > 3:
                # Convert to full URL
                if href.startswith('/portal/docs'):
                    full_url = self.base_url + href
                elif not href.startswith('http'):
                    full_url = self.base_url + '/portal/docs' + href
                else:
                    full_url = href
                
                endpoint_links.append({
                    'text': text,
                    'url': full_url
                })
        
        # Remove duplicates
        seen = set()
        unique_links = []
        for link in endpoint_links:
            if link['url'] not in seen:
                seen.add(link['url'])
                unique_links.append(link)
        
        print(f"✓ Found {len(unique_links)} unique endpoint links\n")
        return unique_links
    
    def extract_table_data(self, table):
        """
        Extract data from a table
        """
        if not table:
            return []
        
        rows = table.find_all('tr')
        data = []
        
        for row in rows[1:]:  # Skip header
            cells = row.find_all(['td', 'th'])
            if len(cells) >= 3:
                data.append({
                    'field': cells[0].get_text(strip=True),
                    'data_type': cells[1].get_text(strip=True),
                    'description': cells[2].get_text(strip=True)
                })
            elif len(cells) == 2:
                data.append({
                    'field': cells[0].get_text(strip=True),
                    'data_type': '',
                    'description': cells[1].get_text(strip=True)
                })
        
        return data
    
    def extract_endpoint_details(self, url, endpoint_text):
        """
        Extract endpoint details from a documentation page
        """
        print(f"  Scraping details...")
        
        endpoint_data = {
            'endpoint_name': endpoint_text,
            'doc_url': url,
            'endpoint': None,
            'method': None,
            'headers': [],
            'parameters': [],
            'response': [],
            'scrape_status': 'failed',  # Will be updated to 'success' if scraping works
            'error_message': None
        }
        
        try:
            iframe_html, iframe_src = self.get_iframe_content(url)
            
            if not iframe_html:
                print(f"  ✗ Could not load iframe")
                endpoint_data['error_message'] = 'Failed to load iframe after retries'
                return endpoint_data
            
            soup = BeautifulSoup(iframe_html, 'html.parser')
            
            # Extract endpoint URL
            text_content = soup.get_text()
            
            # Look for API endpoint URL
            api_url_pattern = r'https://business-api\.tiktok\.com/open_api/[v\d\./a-zA-Z_]+'
            matches = re.findall(api_url_pattern, text_content)
            if matches:
                endpoint_data['endpoint'] = matches[0]
            
            if not endpoint_data['endpoint']:
                rel_url_pattern = r'/open_api/[v\d\./a-zA-Z_]+'
                matches = re.findall(rel_url_pattern, text_content)
                if matches:
                    endpoint_data['endpoint'] = 'https://business-api.tiktok.com' + matches[0]
            
            # Extract HTTP method
            method_keywords = ['POST', 'GET', 'PUT', 'DELETE', 'PATCH']
            for method in method_keywords:
                if re.search(rf'\b{method}\b', text_content, re.IGNORECASE):
                    endpoint_data['method'] = method
                    break
            
            # Extract tables
            all_headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
            
            for heading in all_headings:
                heading_text = heading.get_text(strip=True).lower()
                next_table = heading.find_next('table')
                
                if not next_table:
                    continue
                
                if 'header' in heading_text and not endpoint_data['headers']:
                    endpoint_data['headers'] = self.extract_table_data(next_table)
                    if endpoint_data['headers']:
                        print(f"    ✓ Extracted {len(endpoint_data['headers'])} headers")
                
                elif 'parameter' in heading_text and 'response' not in heading_text and not endpoint_data['parameters']:
                    endpoint_data['parameters'] = self.extract_table_data(next_table)
                    if endpoint_data['parameters']:
                        print(f"    ✓ Extracted {len(endpoint_data['parameters'])} parameters")
                
                elif 'response' in heading_text and not endpoint_data['response']:
                    endpoint_data['response'] = self.extract_table_data(next_table)
                    if endpoint_data['response']:
                        print(f"    ✓ Extracted {len(endpoint_data['response'])} response fields")
            
            print(f"  ✓ Endpoint: {endpoint_data['endpoint']}")
            print(f"  ✓ Method: {endpoint_data['method']}")
            
            endpoint_data['scrape_status'] = 'success'
            
            return endpoint_data
            
        except Exception as e:
            print(f"  ✗ Error: {str(e)[:100]}")
            endpoint_data['error_message'] = str(e)[:200]
            return endpoint_data
    
    def scrape_all_endpoints(self, start_url, delay=2, save_interval=10):
        """
        Main scraping function with progress saving and error recovery
        """
        self.setup_driver()
        
        try:
            # Get endpoint links
            endpoint_links = self.get_endpoint_links_from_page(start_url)
            
            if not endpoint_links:
                print("\n✗ No endpoint links found!")
                return []
            
            print("\n" + "="*80)
            print("STEP 2: Scraping each endpoint's details")
            print("="*80 + "\n")
            
            all_data = []
            failed_count = 0
            success_count = 0
            
            for i, link_info in enumerate(endpoint_links, 1):
                print(f"[{i}/{len(endpoint_links)}] {link_info['text']}")
                
                endpoint_data = self.extract_endpoint_details(link_info['url'], link_info['text'])
                
                # Always append data, even if failed
                all_data.append(endpoint_data)
                
                if endpoint_data['scrape_status'] == 'success':
                    success_count += 1
                else:
                    failed_count += 1
                    print(f"  ⚠ Status: FAILED")
                
                print()
                
                # Save progress periodically
                if i % save_interval == 0:
                    print(f"  💾 Saving progress... ({i}/{len(endpoint_links)} endpoints)")
                    self.save_to_json(all_data, f'tiktok_api_docs_progress_{i}.json')
                    print(f"  ✓ Progress saved")
                    print()
                
                # Add delay between requests
                if i < len(endpoint_links):
                    time.sleep(delay)
            
            # Print final statistics
            print("\n" + "="*80)
            print("SCRAPING COMPLETE")
            print("="*80)
            print(f"✓ Successfully scraped: {success_count}/{len(endpoint_links)}")
            if failed_count > 0:
                print(f"⚠ Failed to scrape: {failed_count}/{len(endpoint_links)}")
            print()
            
            return all_data
            
        finally:
            if self.driver:
                self.driver.quit()
    
    def save_to_json(self, data, filename='tiktok_api_docs.json'):
        """Save to JSON"""
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"\n✓ JSON data saved to {filename}")
    
    def save_to_excel(self, data, filename='tiktok_api_docs.xlsx'):
        """Save to Excel with unified metrics and failure marking"""
        # Create summary with status marking
        summary_data = []
        for item in data:
            summary_data.append({
                'Status': item.get('scrape_status', 'unknown').upper(),
                'Endpoint Name': item.get('endpoint_name', ''),
                'Documentation URL': item.get('doc_url', ''),
                'API Endpoint': item.get('endpoint', 'NOT FOUND'),
                'Method': item.get('method', 'NOT FOUND'),
                'Headers Count': len(item.get('headers', [])),
                'Parameters Count': len(item.get('parameters', [])),
                'Response Fields Count': len(item.get('response', [])),
                'Error Message': item.get('error_message', '')
            })
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write summary sheet
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Apply formatting to highlight failed rows
            workbook = writer.book
            worksheet = writer.sheets['Summary']
            
            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            
            for row_idx, row in enumerate(summary_df.itertuples(), start=2):
                if row.Status == 'FAILED':
                    for col_idx in range(1, len(summary_df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
                elif row.Status == 'SUCCESS':
                    worksheet.cell(row=row_idx, column=1).fill = green_fill
            
            # Create unified metrics sheets (only for successful scrapes)
            successful_data = [item for item in data if item.get('scrape_status') == 'success']
            if successful_data:
                self._create_unified_metrics_sheets(successful_data, writer)
            
            # Individual endpoint sheets
            for i, item in enumerate(data, 1):
                sheet_name = item.get('endpoint_name', f'Endpoint_{i}')[:31]
                
                details_data = []
                details_data.append(['Status', item.get('scrape_status', 'unknown').upper()])
                details_data.append(['Endpoint Name', item.get('endpoint_name', '')])
                details_data.append(['Documentation URL', item.get('doc_url', '')])
                details_data.append(['API Endpoint', item.get('endpoint', 'NOT FOUND')])
                details_data.append(['Method', item.get('method', 'NOT FOUND')])
                
                if item.get('error_message'):
                    details_data.append(['Error Message', item.get('error_message', '')])
                
                details_data.append(['', ''])
                
                if item.get('headers'):
                    details_data.append(['HEADERS', '', ''])
                    details_data.append(['Field', 'Data Type', 'Description'])
                    for h in item['headers']:
                        details_data.append([h.get('field', ''), h.get('data_type', ''), h.get('description', '')])
                    details_data.append(['', '', ''])
                
                if item.get('parameters'):
                    details_data.append(['PARAMETERS', '', ''])
                    details_data.append(['Field', 'Data Type', 'Description'])
                    for p in item['parameters']:
                        details_data.append([p.get('field', ''), p.get('data_type', ''), p.get('description', '')])
                    details_data.append(['', '', ''])
                
                if item.get('response'):
                    details_data.append(['RESPONSE', '', ''])
                    details_data.append(['Field', 'Data Type', 'Description'])
                    for r in item['response']:
                        details_data.append([r.get('field', ''), r.get('data_type', ''), r.get('description', '')])
                
                df = pd.DataFrame(details_data)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        print(f"✓ Excel data saved to {filename}")
        
        # Print failure summary
        failed_endpoints = [item for item in data if item.get('scrape_status') == 'failed']
        if failed_endpoints:
            print(f"\n⚠ Warning: {len(failed_endpoints)} endpoints failed to scrape:")
            for item in failed_endpoints[:10]:  # Show first 10
                print(f"  - {item['endpoint_name']}")
            if len(failed_endpoints) > 10:
                print(f"  ... and {len(failed_endpoints) - 10} more (see Summary sheet)")
            print(f"\n  Failed endpoints are highlighted in RED in the Summary sheet")
    
    def _create_unified_metrics_sheets(self, data, writer):
        """Create unified metrics sheets"""
        # [Same implementation as in the requests-based scraper]
        # Collect unique fields
        headers_map = {}
        parameters_map = {}
        response_map = {}
        
        for item in data:
            endpoint_info = {
                'endpoint_name': item.get('endpoint_name', ''),
                'api_endpoint': item.get('endpoint', ''),
                'doc_url': item.get('doc_url', ''),
                'method': item.get('method', '')
            }
            
            for header in item.get('headers', []):
                field = header.get('field', '')
                if field:
                    if field not in headers_map:
                        headers_map[field] = {
                            'data_type': header.get('data_type', ''),
                            'description': header.get('description', ''),
                            'endpoints': []
                        }
                    headers_map[field]['endpoints'].append(endpoint_info)
            
            for param in item.get('parameters', []):
                field = param.get('field', '')
                if field:
                    if field not in parameters_map:
                        parameters_map[field] = {
                            'data_type': param.get('data_type', ''),
                            'description': param.get('description', ''),
                            'endpoints': []
                        }
                    parameters_map[field]['endpoints'].append(endpoint_info)
            
            for resp in item.get('response', []):
                field = resp.get('field', '')
                if field:
                    if field not in response_map:
                        response_map[field] = {
                            'data_type': resp.get('data_type', ''),
                            'description': resp.get('description', ''),
                            'endpoints': []
                        }
                    response_map[field]['endpoints'].append(endpoint_info)
        
        # Create sheets (same format as before)
        for field_map, sheet_name in [
            (headers_map, 'All Headers'),
            (parameters_map, 'All Parameters'),
            (response_map, 'All Response Fields')
        ]:
            if field_map:
                rows = []
                for field, info in sorted(field_map.items()):
                    for idx, ep in enumerate(info['endpoints']):
                        rows.append({
                            'Field': field if idx == 0 else '',
                            'Data Type': info['data_type'] if idx == 0 else '',
                            'Description': info['description'] if idx == 0 else '',
                            'Usage Count': len(info['endpoints']) if idx == 0 else '',
                            'Endpoint Name': ep['endpoint_name'],
                            'API Endpoint URL': ep['api_endpoint'],
                            'Method': ep['method'],
                            'Documentation URL': ep['doc_url']
                        })
                    rows.append({k: '' for k in rows[0].keys()})  # Empty row
                
                pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"  ✓ Created unified metrics sheets")


def main():
    print("\n" + "="*80)
    print("TikTok Business API Documentation Scraper (Selenium)")
    print("="*80 + "\n")
    
    scraper = TikTokAPIDocScraperSelenium(headless=True)
    start_url = "https://business-api.tiktok.com/portal/docs?id=1735713875563521"
    
    # Scrape with progress saving every 10 endpoints
    data = scraper.scrape_all_endpoints(start_url, delay=2, save_interval=10)
    
    if data:
        # Calculate statistics
        total = len(data)
        successful = len([d for d in data if d.get('scrape_status') == 'success'])
        failed = len([d for d in data if d.get('scrape_status') == 'failed'])
        
        print("\n" + "="*80)
        print(f"RESULTS: Scraped {total} endpoints")
        print("="*80)
        print(f"✓ Successful: {successful}")
        if failed > 0:
            print(f"⚠ Failed: {failed}")
        print()
        
        print("Saving data...")
        
        scraper.save_to_json(data)
        scraper.save_to_excel(data)
        
        print("\n" + "="*80)
        print("OUTPUT FILES")
        print("="*80)
        print("✓ tiktok_api_docs.json - Complete data (includes failed attempts)")
        print("✓ tiktok_api_docs.xlsx - Excel with multiple sheets:")
        print("  - Summary: Overview with status indicators")
        print("    • GREEN = Successfully scraped")
        print("    • RED = Failed to scrape")
        print("  - All Headers: Unified list (successful endpoints only)")
        print("  - All Parameters: Unified list (successful endpoints only)")
        print("  - All Response Fields: Unified list (successful endpoints only)")
        print("  - Individual endpoint sheets with full details")
        
        if failed > 0:
            print(f"\n⚠ Note: {failed} endpoints failed to scrape.")
            print("  Check the Summary sheet for details.")
            print("  Failed endpoints are highlighted in RED.")
            print("  You can retry these manually or run the scraper again.")
    else:
        print("\n✗ No data was scraped.")


if __name__ == "__main__":
    main()